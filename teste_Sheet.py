# Conexão Sheets
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# Planilhas
from openpyxl import load_workbook
import pandas as pd

# Arquivos
import os
from pathlib import Path

# CONEXÃO COM GOOGLE SHEETS
filename = "ARQUIVO_G_SHEET" # ARQUIVO GERADO NO GOOGLE CLOUD PELAS CONTAS DE SERVIÇO

    # PADRÃO
scopes = [
    "https://spreadsheets.google.com/feeds",
    "https://www.googleapis.com/auth/drive"
]

    # PADRÃO
creds = ServiceAccountCredentials.from_json_keyfile_name(filename=filename,scopes=scopes)
client = gspread.authorize(creds)


titulo = "NOME_PLANILHA" # NOME DO SHEET
folder_id = "ID_PASTA_DRIVE" # ID DA PASTA NO GOOGLE DRIVE

    # PADRÃO
planilhaCompleta = client.open(title = titulo, folder_id = folder_id)

# NOME DAS PLANILHAS
origem = r"C:\Users\Padrão\Desktop\BeL Tech\VJ\Royal\Planilhas"

for caminho, subpasta, arquivos in os.walk(origem):
    for contador,(nome) in enumerate(arquivos): # UMA EXECUÇÃO POR ARQUIVO
        print(contador)
        arq = caminho+"\\"+nome
        if contador == 0: # SE FOR A PRIMEIRA EXECUÇÃO IMPORTA PRA RESERTAR A PLANILHA
            df_planilha = pd.read_excel(arq) # ABRE PRA CONVERTER EM CSV
            nomeArq = nome+".csv"
            arq = caminho+"\\"+nomeArq
            df_planilha.to_csv(arq) # CONVERTE EM CSV
            content = open(arq, 'r',encoding="utf8").read() # PEGA OS DADOS
            client.import_csv(planilhaCompleta.id,content) # IMPORTA
            deletar = Path(arq) # SELECIONA A PLANILHA À APAGAR
            deletar.unlink() # APAGA A PLANILHA
        else: # CASO NÃO FOR A PRIMEIRA EXECUÇÃO
            arquivo = load_workbook(arq) # SELECIONA OS DADOS
            aba_atual = arquivo.active # PEGA O NOME DA ABA
            data = [] # VARIÁVEL UTILIZADA PARA GUARDAR OS DADOS DAS PLANILHAS
            aux = [] # AUXILIAR PRA SALVAR LINHAS NA VARIÁVEL "data"
            planilhaCompleta.add_worksheet(nome,200,100,index=None) # ADICIONA UMA PÁGINA AO SHEET
            for i in range(1,aba_atual.max_row+1): # 2 FOR PRA SALVAR OS DADOS DA PLANILHA LINHA A LINHA
                for y in range(1,aba_atual.max_column+1):
                    if str(aba_atual.cell(row=i, column=y).value) == "None":
                        aux.append("")
                    else:
                        aux.append(str(aba_atual.cell(row=i, column=y).value))

                data.append(aux)
                aux = []

            # INSERE OS DADOS NA ABA CORRETA
            planilhaCompleta.values_update(
                nome,
                params={
                    'valueInputOption': 'USER_ENTERED'
                    },
                body={
                    'values':data
                }
            )
